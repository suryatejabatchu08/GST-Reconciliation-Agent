"""
report_service/tests/test_report_builders.py
Unit tests for PDF and Excel report builders.
Tests generate real files in memory — no file system or DB dependency.
"""

import pytest
from report_service.pdf_builder import build_pdf_report, fmt_inr as pdf_fmt_inr
from report_service.excel_builder import build_excel_report


# ── Shared test fixtures ───────────────────────────────────

def sample_escalation():
    return {
        "invoice_no": "INV-ESC-001",
        "supplier_name": "Risky Supplier Pvt Ltd",
        "gstin": "29AABCT1332L1ZT",
        "mismatch_type": "missing",
        "severity": "escalate",
        "itc_risk_amount": 45000.0,
        "cause_reasoning": "Invoice not found in GSTR-2A. Supplier has not filed.",
        "recommended_action": "Contact supplier and request GSTR-1 amendment.",
    }

def sample_followup():
    return {
        "invoice_no": "INV-FUP-001",
        "supplier_name": "Small Vendor Co",
        "gstin": "27AAKCS9175D1Z5",
        "mismatch_type": "amount",
        "severity": "followup",
        "itc_risk_amount": 900.0,
        "cause_reasoning": "Amount in books ₹59,000 vs portal ₹58,100.",
        "recommended_action": "Request corrected invoice.",
    }

def sample_auto():
    return {
        "invoice_no": "INV-AUTO-001",
        "supplier_name": "Rounding Vendor",
        "gstin": "07AAACH7409R1ZZ",
        "mismatch_type": "amount",
        "severity": "auto",
        "itc_risk_amount": 0.50,
        "cause_reasoning": "Rounding difference of ₹0.50.",
        "recommended_action": "Auto-resolved.",
    }

REPORT_KWARGS = dict(
    client_name="Test Client Pvt Ltd",
    client_gstin="29ABCDE1234F1Z5",
    filing_period="2024-03",
    ca_name="CA Sharma",
    firm_name="Sharma & Associates",
    total_invoices=100,
    matched_count=90,
    auto_mismatches=[sample_auto()],
    followup_mismatches=[sample_followup()],
    escalation_mismatches=[sample_escalation()],
    total_itc_at_risk=45900.0,
)


class TestPDFBuilder:

    def test_returns_bytes(self):
        result = build_pdf_report(**REPORT_KWARGS)
        assert isinstance(result, bytes)

    def test_output_is_nonempty(self):
        result = build_pdf_report(**REPORT_KWARGS)
        assert len(result) > 5000   # A real PDF is always > 5 KB

    def test_starts_with_pdf_magic_bytes(self):
        """PDF files always start with %PDF"""
        result = build_pdf_report(**REPORT_KWARGS)
        assert result[:4] == b"%PDF"

    def test_no_mismatches_still_generates(self):
        """Report with no mismatches should still produce a valid PDF."""
        kwargs = {**REPORT_KWARGS,
                  "auto_mismatches": [],
                  "followup_mismatches": [],
                  "escalation_mismatches": [],
                  "total_itc_at_risk": 0.0}
        result = build_pdf_report(**kwargs)
        assert result[:4] == b"%PDF"

    def test_fmt_inr_used_in_pdf(self):
        """Verify fmt_inr works correctly for PDF amounts."""
        assert pdf_fmt_inr(45000.0) == "₹45,000.00"
        assert pdf_fmt_inr(0) == "₹0.00"
        assert pdf_fmt_inr(150000.0) == "₹1,50,000.00"


class TestExcelBuilder:

    def test_returns_bytes(self):
        result = build_excel_report(**REPORT_KWARGS)
        assert isinstance(result, bytes)

    def test_output_is_nonempty(self):
        result = build_excel_report(**REPORT_KWARGS)
        assert len(result) > 2000

    def test_starts_with_xlsx_magic_bytes(self):
        """XLSX files are ZIP archives starting with PK magic bytes."""
        result = build_excel_report(**REPORT_KWARGS)
        assert result[:2] == b"PK"

    def test_no_mismatches_still_generates(self):
        """Excel with no mismatches should still produce valid file."""
        kwargs = {**REPORT_KWARGS,
                  "auto_mismatches": [],
                  "followup_mismatches": [],
                  "escalation_mismatches": [],
                  "total_itc_at_risk": 0.0}
        result = build_excel_report(**kwargs)
        assert result[:2] == b"PK"

    def test_excel_can_be_loaded_by_openpyxl(self):
        """Generated Excel should be a valid workbook openpyxl can open."""
        import io
        from openpyxl import load_workbook
        result = build_excel_report(**REPORT_KWARGS)
        wb = load_workbook(io.BytesIO(result))
        assert "Summary" in wb.sheetnames

    def test_excel_has_correct_sheets(self):
        """Should have Summary + at least the mismatch sheet."""
        import io
        from openpyxl import load_workbook
        result = build_excel_report(**REPORT_KWARGS)
        wb = load_workbook(io.BytesIO(result))
        assert "Summary" in wb.sheetnames
        assert "All Mismatches" in wb.sheetnames

    def test_summary_sheet_has_data(self):
        """Summary sheet should have at least 5 rows of data."""
        import io
        from openpyxl import load_workbook
        result = build_excel_report(**REPORT_KWARGS)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Summary"]
        assert ws.max_row >= 5
