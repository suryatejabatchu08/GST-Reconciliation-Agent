"""
report_service/excel_builder.py
Generates an Excel (.xlsx) reconciliation export using openpyxl.

Sheets:
  1. Summary       — same KPIs as the PDF cover, formatted table
  2. All Mismatches — full mismatch table (sortable/filterable in Excel)
  3. Escalations   — escalation-only sheet for the CA to work through
  4. Follow-ups    — follow-up sheet for supplier tracking

Design:
  - Header row frozen + auto-filter on all data sheets
  - Conditional formatting: red fill for escalate rows, amber for followup
  - Column widths auto-adjusted based on content
  - Indian number format applied to all amount columns
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Colour constants (openpyxl uses ARGB hex strings) ─────
NAVY_FILL    = PatternFill("solid", fgColor="1a2744")
BLUE_FILL    = PatternFill("solid", fgColor="2563eb")
RED_FILL     = PatternFill("solid", fgColor="FFEDED")
AMBER_FILL   = PatternFill("solid", fgColor="FFFBEB")
GREEN_FILL   = PatternFill("solid", fgColor="F0FDF4")
GREY_FILL    = PatternFill("solid", fgColor="F8FAFC")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Calibri", size=9)
TITLE_FONT   = Font(name="Calibri", bold=True, color="1a2744", size=14)
AMOUNT_FONT  = Font(name="Calibri", size=9)

THIN_BORDER  = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

INR_FORMAT = '₹#,##0.00'    # Excel number format for Indian currency
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")


def _set_header_row(ws, row: int, headers: list[str], fill=BLUE_FILL):
    """Write a styled header row."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _auto_width(ws, min_width=12, max_width=45):
    """Adjust column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _write_mismatch_rows(ws, mismatches: list[dict], start_row: int):
    """Write mismatch rows to a worksheet starting at start_row."""
    severity_fills = {
        "escalate": RED_FILL,
        "followup": AMBER_FILL,
        "auto": GREEN_FILL,
    }

    for i, m in enumerate(mismatches):
        row = start_row + i
        severity = m.get("severity", "")
        fill = severity_fills.get(severity, WHITE_FILL)

        values = [
            i + 1,
            m.get("invoice_no") or "N/A",
            m.get("supplier_name") or m.get("gstin") or "N/A",
            m.get("mismatch_type", "").upper(),
            severity.upper(),
            m.get("itc_risk_amount", 0),
            m.get("cause_reasoning", ""),
            m.get("recommended_action", ""),
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER

            if col == 6:   # ITC Risk column — number format
                cell.number_format = INR_FORMAT
                cell.alignment = RIGHT_ALIGN
            else:
                cell.alignment = LEFT


def build_excel_report(
    client_name: str,
    client_gstin: str,
    filing_period: str,
    ca_name: str,
    firm_name: str,
    total_invoices: int,
    matched_count: int,
    auto_mismatches: list[dict],
    followup_mismatches: list[dict],
    escalation_mismatches: list[dict],
    total_itc_at_risk: float,
) -> bytes:
    """
    Generate an Excel reconciliation report and return it as bytes.
    """
    wb = Workbook()
    all_mismatches = escalation_mismatches + followup_mismatches + auto_mismatches

    # ── Sheet 1: Summary ───────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_view.showGridLines = False

    # Title block
    ws_summary.merge_cells("A1:F1")
    title_cell = ws_summary["A1"]
    title_cell.value = f"GST Reconciliation Report — {client_name}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    title_cell.fill = GREY_FILL

    ws_summary.merge_cells("A2:F2")
    sub_cell = ws_summary["A2"]
    sub_cell.value = f"GSTIN: {client_gstin} | Period: {filing_period} | Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    sub_cell.font = Font(name="Calibri", size=9, italic=True, color="475569")
    sub_cell.alignment = CENTER

    ws_summary.append([])  # blank row

    # KPI table
    kpi_headers = ["Metric", "Value", "Notes"]
    _set_header_row(ws_summary, 4, kpi_headers, fill=NAVY_FILL)

    kpi_data = [
        ("Total Invoices Analysed", total_invoices, "Tally / Zoho Books + GST Portal"),
        ("Clean Matches", matched_count, "Fully reconciled"),
        ("Auto-Resolved (Rounding)", len(auto_mismatches), "Difference ≤ ₹1"),
        ("Needs Supplier Follow-up", len(followup_mismatches), "Supplier notified by email"),
        ("Requires CA Review (Escalations)", len(escalation_mismatches), "High ITC risk"),
        ("Total Mismatches", len(all_mismatches), ""),
        ("Total ITC at Risk (₹)", total_itc_at_risk, "Across all unresolved mismatches"),
        ("Prepared by", f"{ca_name} — {firm_name}", ""),
    ]

    for row_idx, (metric, value, note) in enumerate(kpi_data, start=5):
        fill = RED_FILL if metric.startswith("Requires") else WHITE_FILL if row_idx % 2 else GREY_FILL
        for col_idx, val in enumerate([metric, value, note], start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if col_idx == 2 and isinstance(val, float):
                cell.number_format = INR_FORMAT
                cell.alignment = RIGHT_ALIGN
            else:
                cell.alignment = LEFT

    ws_summary.column_dimensions["A"].width = 38
    ws_summary.column_dimensions["B"].width = 18
    ws_summary.column_dimensions["C"].width = 30
    ws_summary.row_dimensions[1].height = 22

    # ── Sheet 2: All Mismatches ────────────────────────────
    if all_mismatches:
        ws_all = wb.create_sheet("All Mismatches")
        ws_all.sheet_view.showGridLines = False

        headers = ["#", "Invoice No", "Supplier", "Type", "Severity", "ITC Risk (₹)", "Cause", "Recommended Action"]
        _set_header_row(ws_all, 1, headers)
        ws_all.freeze_panes = "A2"
        ws_all.auto_filter.ref = f"A1:H{len(all_mismatches)+1}"

        _write_mismatch_rows(ws_all, all_mismatches, start_row=2)
        _auto_width(ws_all)
        ws_all.column_dimensions["G"].width = 50
        ws_all.column_dimensions["H"].width = 45

    # ── Sheet 3: Escalations only ──────────────────────────
    if escalation_mismatches:
        ws_esc = wb.create_sheet("🔴 Escalations")
        ws_esc.sheet_view.showGridLines = False
        ws_esc.tab_color = "DC2626"

        headers = ["#", "Invoice No", "Supplier", "Type", "ITC Risk (₹)", "Cause", "Recommended Action"]
        _set_header_row(ws_esc, 1, headers, fill=PatternFill("solid", fgColor="DC2626"))
        ws_esc.freeze_panes = "A2"
        ws_esc.auto_filter.ref = f"A1:G{len(escalation_mismatches)+1}"

        for i, m in enumerate(escalation_mismatches):
            row = i + 2
            values = [
                i + 1,
                m.get("invoice_no") or "N/A",
                m.get("supplier_name") or m.get("gstin") or "N/A",
                m.get("mismatch_type", "").upper(),
                m.get("itc_risk_amount", 0),
                m.get("cause_reasoning", ""),
                m.get("recommended_action", ""),
            ]
            for col, val in enumerate(values, start=1):
                cell = ws_esc.cell(row=row, column=col, value=val)
                cell.font = BODY_FONT
                cell.fill = RED_FILL
                cell.border = THIN_BORDER
                cell.alignment = RIGHT_ALIGN if col == 5 else LEFT
                if col == 5:
                    cell.number_format = INR_FORMAT

        _auto_width(ws_esc)

    # ── Sheet 4: Follow-ups ────────────────────────────────
    if followup_mismatches:
        ws_fup = wb.create_sheet("🟡 Follow-ups")
        ws_fup.sheet_view.showGridLines = False
        ws_fup.tab_color = "D97706"

        headers = ["#", "Invoice No", "Supplier", "Type", "ITC Risk (₹)", "Cause", "Action Taken"]
        _set_header_row(ws_fup, 1, headers, fill=PatternFill("solid", fgColor="D97706"))
        ws_fup.freeze_panes = "A2"

        for i, m in enumerate(followup_mismatches):
            row = i + 2
            values = [
                i + 1,
                m.get("invoice_no") or "N/A",
                m.get("supplier_name") or m.get("gstin") or "N/A",
                m.get("mismatch_type", "").upper(),
                m.get("itc_risk_amount", 0),
                m.get("cause_reasoning", ""),
                "Email sent to supplier",
            ]
            for col, val in enumerate(values, start=1):
                cell = ws_fup.cell(row=row, column=col, value=val)
                cell.font = BODY_FONT
                cell.fill = AMBER_FILL
                cell.border = THIN_BORDER
                cell.alignment = RIGHT_ALIGN if col == 5 else LEFT
                if col == 5:
                    cell.number_format = INR_FORMAT

        _auto_width(ws_fup)

    # ── Save to bytes ──────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
